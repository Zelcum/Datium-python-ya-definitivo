from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status, exceptions
from django.shortcuts import get_object_or_404
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from datetime import timedelta
import uuid, os, csv, json, logging, time
from io import StringIO, BytesIO
from django.contrib.auth.hashers import make_password, check_password

from .models import (
    Plan, User, System, SystemTable, SystemField,
    SystemFieldOption, SystemRecord, SystemRecordValue,
    SystemRelationship, AuditLog, SecurityAudit,
    SystemCollaborator, SystemCollaboratorTable, AppSetting, UserReport, BlockedIP,
    Discount, Payment
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import validators as val
from .plans import (
    sync_plans_into_db,
    assert_can_create_system,
    assert_can_add_table,
    plan_for_user,
    user_has_ai,
)
from .security_locks import (
    is_login_locked,
    record_login_failure,
    clear_login_failures,
    is_password_verify_locked,
    record_password_verify_failure,
    clear_password_verify_failures,
)

logger = logging.getLogger(__name__)
ADMIN_EMAIL = os.getenv('DATIUM_ADMIN_EMAIL', 'ibzantrabajo@gmail.com').lower()
ADMIN_PASSWORD = os.getenv('DATIUM_ADMIN_PASSWORD', 'Datium777')


def _verify_password(plain, hashed):
    if not hashed:
        return False
    if hashed.startswith(('pbkdf2_', 'bcrypt', 'argon2', 'scrypt')):
        return check_password(plain, hashed)
    return plain == hashed


def _ensure_hashed(user):
    if not user.password_hash.startswith(('pbkdf2_', 'bcrypt', 'argon2', 'scrypt')):
        user.password_hash = make_password(user.password_hash)
        user.save(update_fields=['password_hash'])


def get_current_user(request):
    uid = request.session.get('user_id')
    if not uid:
        return None
    try:
        user = User.objects.get(id=uid)
        if user.is_suspended:
            return None
        return user
    except User.DoesNotExist:
        return None

def seed_data():
    sync_plans_into_db()
    plan_corp = Plan.objects.filter(name='Corporate').first()
    admin_user, created = User.objects.get_or_create(
        email=ADMIN_EMAIL,
        defaults={
            'name': 'Administrador Datium',
            'password_hash': make_password(ADMIN_PASSWORD),
            'role': 'admin',
            'plan': plan_corp
        }
    )
    if not created and not admin_user.password_hash.startswith(('pbkdf2_', 'bcrypt', 'argon2')):
        admin_user.password_hash = make_password(ADMIN_PASSWORD)
        admin_user.role = 'admin'
        admin_user.save()
    User.objects.exclude(email__iexact=ADMIN_EMAIL).filter(role='admin').update(role='user')


def require_auth(request):
    user = get_current_user(request)
    if not user:
        return None, Response({'error': 'No autenticado'}, status=status.HTTP_401_UNAUTHORIZED)
    return user, None

def require_admin(request):
    user = get_current_user(request)
    if not user:
        return None, Response({'error': 'No autenticado'}, status=401)
    if user.email.lower() != ADMIN_EMAIL:
        return None, Response({'error': 'No tienes permisos de administrador global.'}, status=403)
    return user, None


def check_system_permission(user, system_id, permission_type):
    """
    Verifica si un usuario tiene un permiso específico sobre un sistema.
    permission_type: 'read', 'create', 'update', 'delete'
    """
    try:
        sys = System.objects.get(id=system_id)
    except System.DoesNotExist:
        return False, Response({'error': 'Sistema no encontrado'}, status=404)

    perms = get_system_permissions(user, sys)
    if perms.get(permission_type):
        return True, None
    return False, Response({'error': f'No tienes permiso de {permission_type} en este sistema'}, status=403)


def get_system_permissions(user, system):
    """Calcula el objeto de permisos para un usuario y sistema dado."""
    if system.owner_id == user.id or user.role == 'admin':
        return {'read': True, 'create': True, 'update': True, 'delete': True, 'is_owner': True}
    try:
        collab = SystemCollaborator.objects.get(system=system, user=user)
        return {
            'read': collab.can_read,
            'create': collab.can_create,
            'update': False,
            'delete': False,
            'is_owner': False,
        }
    except SystemCollaborator.DoesNotExist:
        return {'read': False, 'create': False, 'update': False, 'delete': False, 'is_owner': False}


def collaborator_table_perm(collab, table, perm):
    fid = {'read': 'can_read', 'create': 'can_create', 'update': 'can_update', 'delete': 'can_delete'}[perm]
    rows = SystemCollaboratorTable.objects.filter(collaborator=collab)
    if rows.exists():
        row = rows.filter(table=table).first()
        if not row:
            return False
        return bool(getattr(row, fid))
    return bool(getattr(collab, fid))


def tables_queryset_for_user(user, system_id):
    try:
        system = System.objects.get(id=system_id)
    except System.DoesNotExist:
        return SystemTable.objects.none()
    if system.owner_id == user.id or user.role == 'admin':
        return SystemTable.objects.filter(system_id=system_id, is_deleted=False).order_by('name')
    collab = SystemCollaborator.objects.filter(system_id=system_id, user=user).first()
    if not collab or not collab.can_read:
        return SystemTable.objects.none()
    rows = SystemCollaboratorTable.objects.filter(collaborator=collab)
    if not rows.exists():
        return SystemTable.objects.filter(system_id=system_id, is_deleted=False).order_by('name')
    return SystemTable.objects.filter(
        system_id=system_id,
        is_deleted=False,
        id__in=rows.filter(can_read=True).values_list('table_id', flat=True),
    ).order_by('name')


def get_table_permissions(user, table):
    system = table.system
    if system.owner_id == user.id or user.role == 'admin':
        return {'read': True, 'create': True, 'update': True, 'delete': True}
    try:
        collab = SystemCollaborator.objects.get(system=system, user=user)
    except SystemCollaborator.DoesNotExist:
        return {'read': False, 'create': False, 'update': False, 'delete': False}
    return {
        'read': collaborator_table_perm(collab, table, 'read'),
        'create': collaborator_table_perm(collab, table, 'create'),
        'update': collaborator_table_perm(collab, table, 'update'),
        'delete': collaborator_table_perm(collab, table, 'delete'),
    }


def check_table_permission(user, table, permission_type):
    if get_table_permissions(user, table).get(permission_type):
        return True, None
    return False, Response({'error': f'Sin permiso de {permission_type} en esta tabla'}, status=403)


def serialize_system(s, user=None):
    d = {
        'id': s.id, 'name': s.name, 'description': s.description or '',
        'imageUrl': s.image_url or '', 'ownerId': s.owner_id,
        'securityMode': s.security_mode or 'none',
        'createdAt': s.created_at.isoformat() if s.created_at else None,
        'userCount': 1,
    }
    if user is not None:
        d['permissions'] = get_system_permissions(user, s)
        d['isOwner'] = s.owner_id == user.id
        d['hasAi'] = user_has_ai(user)
    return d


def serialize_table(t, user=None):
    res = {
        'id': t.id, 'name': t.name, 'description': t.description or '',
        'systemId': t.system_id,
        'createdAt': t.created_at.isoformat() if t.created_at else None,
    }
    if user is not None:
        res['permissions'] = get_table_permissions(user, t)
    return res


def serialize_field(f):
    r = {
        'id': f.id, 'name': f.name, 'type': f.type,
        'required': f.required, 'is_unique': f.is_unique,
        'isPrimaryKey': f.is_primary_key, 'orderIndex': f.order_index,
    }
    if f.type == 'select':
        r['options'] = list(SystemFieldOption.objects.filter(field=f).values_list('value', flat=True))
    if f.type == 'relation' and f.related_table_id:
        r['relatedTableId'] = f.related_table_id
        if f.related_display_field_id:
            r['relatedDisplayFieldId'] = f.related_display_field_id
            try:
                r['relatedFieldName'] = f.related_display_field.name
            except Exception:
                r['relatedFieldName'] = None
        else:
            r['relatedFieldName'] = None
            r['relatedDisplayFieldId'] = None
    return r


def serialize_record(record, fields):
    vals = {v.field_id: v.value for v in SystemRecordValue.objects.filter(record=record)}
    fv = {}
    display_vals = {}
    for f in fields:
        val = vals.get(f.id, '')
        fv[f.name] = val
        
        # Resolve relations for display
        if f.type == 'relation' and val:
            try:
                rel_id = int(val)
                rel_rec = SystemRecord.objects.get(id=rel_id)
                disp_field = f.related_display_field
                if not disp_field and f.related_table_id:
                    # Fallback to PK of related table
                    disp_field = SystemField.objects.filter(table_id=f.related_table_id, is_primary_key=True).first()
                
                if disp_field:
                    # Get the value of the display field in the related record
                    d_val = SystemRecordValue.objects.filter(record=rel_rec, field=disp_field).first()
                    display_vals[f.name] = d_val.value if d_val else f"#{rel_id}"
                else:
                    # Total fallback to first field's value if no PK or display field
                    first_val = SystemRecordValue.objects.filter(record=rel_rec).exclude(field__type='file').first()
                    display_vals[f.name] = first_val.value if first_val else f"#{rel_id}"
            except Exception:
                display_vals[f.name] = val
        else:
            display_vals[f.name] = val
            
    return {'id': record.id, 'fieldValues': fv, 'displayValues': display_vals}
 
 
def _validate_record_values(table, fields, values, record_id=None):
    """
    Validates values for a record in a table.
    - Required fields
    - Uniqueness (is_unique or is_primary_key)
    - Relation existence
    """
    for f in fields:
        val = values.get(str(f.id))
        if val is None:
            val = values.get(f.name)
        
        # Required check
        if f.required and (val is None or str(val).strip() == ""):
            return f'El campo "{f.name}" es obligatorio.'
        
        if val is not None and str(val).strip() != "":
            # Uniqueness check
            if f.is_unique or f.is_primary_key:
                # Find other records with the same value for this field
                others = SystemRecordValue.objects.filter(field=f, value=val).exclude(record_id=record_id)
                if others.exists():
                    return f'El valor "{val}" ya existe en el campo único "{f.name}".'
            
            # Relation check
            if f.type == 'relation' and f.related_table_id:
                try:
                    rel_id = int(val)
                    if not SystemRecord.objects.filter(id=rel_id, table_id=f.related_table_id).exists():
                        return f'El registro relacionado con ID {rel_id} no existe en la tabla "{f.related_table.name}".'
                except (ValueError, TypeError):
                    return f'El valor del campo "{f.name}" debe ser un ID numérico válido.'
    return None


def log_action(user, system, action, details='', ip=''):
    AuditLog.objects.create(user=user, system=system, action=action, details=details, ip=ip or '')


# ═══════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════

@api_view(['POST'])
def login_view(request):
    seed_data()
    email = request.data.get('email', '').strip()
    password = request.data.get('password', '')
    ip = request.META.get('REMOTE_ADDR', '') or 'unknown'
    lock_id = email.lower() or ip

    if is_login_locked(lock_id):
        return Response(
            {'error': 'Demasiados intentos fallidos. Espera unos minutos e intenta de nuevo.'},
            status=status.HTTP_429_TOO_MANY_REQUESTS,
        )

    try:
        user = User.objects.filter(email__iexact=email).first()
        if user and _verify_password(password, user.password_hash):
            _ensure_hashed(user)
            clear_login_failures(lock_id)
            request.session['user_id'] = user.id
            return Response({
                'token': str(uuid.uuid4()),
                'usuario': {
                    'id': user.id, 'name': user.name, 'email': user.email,
                    'avatarUrl': user.avatar_url or '',
                    'role': user.role,
                    'terms_version_accepted': user.terms_version_accepted,
                    'session_timeout_minutes': user.session_timeout_minutes,
                }
            })
        record_login_failure(lock_id)
        return Response({'error': 'Credenciales inválidas'}, status=401)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def registro_view(request):
    try:
        seed_data()
        nombre = (request.data.get('nombre') or '').strip()
        email = (request.data.get('email') or '').strip().lower()
        password = request.data.get('password', '')
        phone = (request.data.get('phone') or '').strip()
        plan_id = request.data.get('planId', 1)

        if not email or not nombre or not password:
            return Response({'error': 'Nombre, email y password son obligatorios'}, status=400)

        err_email = val.validate_email(email)
        if err_email:
            return Response({'error': err_email}, status=400)

        err = val.validate_profile_name(nombre)
        if err:
            return Response({'error': err}, status=400)

        if User.objects.filter(email__iexact=email).exists():
            return Response({'error': 'El email ya esta en uso'}, status=400)

        # Unique phone check (if provided)
        if phone and User.objects.filter(phone=phone).exists():
            return Response({'error': 'El telefono ya esta registrado en otra cuenta'}, status=400)

        errp = val.validate_profile_phone(phone)
        if errp:
            return Response({'error': errp}, status=400)

        # Resolve plan
        plan_map = {1: 'Free', '1': 'Free', 2: 'Pro', '2': 'Pro', 3: 'Corporate', '3': 'Corporate'}
        plan_name = plan_map.get(plan_id) or plan_map.get(str(plan_id), 'Free')
        plan = Plan.objects.filter(name__iexact=plan_name).first()

        err_pw = val.validate_new_password(password)
        if err_pw:
            return Response({'error': err_pw}, status=400)

        user = User.objects.create(
            name=nombre,
            email=email,
            phone=phone if phone else None,
            password_hash=make_password(password),
            plan=plan,
            avatar_url=None
        )
        request.session['user_id'] = user.id

        return Response({
            'token': str(uuid.uuid4()),
            'usuario': {'id': user.id, 'name': user.name, 'email': user.email, 'avatarUrl': ''}
        }, status=201)
    except Exception as e:
        logging.error(f"Error en registro: {str(e)}", exc_info=True)
        return Response({'error': f'Error interno: {str(e)}'}, status=500)


@api_view(['POST'])
def logout_view(request):
    request.session.flush()
    return Response({'ok': True})


@api_view(['POST'])
def recuperar_password_view(request):
    return Response({'mensaje': 'Si el email existe, se envió un correo de recuperación.'})


# ═══════════════════════════════════════════
# USER PROFILE
# ═══════════════════════════════════════════

@api_view(['GET', 'PUT'])
def user_profile_view(request):
    user, err = require_auth(request)
    if err:
        return err

    if request.method == 'GET':
        p = plan_for_user(user)
        plan_name = p.name if p else 'Gratuito'
        
        # Terms check
        version_setting = AppSetting.objects.filter(key='terms_version').first()
        current_terms_version = int(version_setting.value if version_setting else 1)
        needs_terms_acceptance = user.terms_version_accepted < current_terms_version
        
        terms_content = ""
        if needs_terms_acceptance:
            content_setting = AppSetting.objects.filter(key='terms_content').first()
            terms_content = content_setting.value if content_setting else "<h1>Términos y Condiciones</h1><p>Por favor acepta para continuar.</p>"

        plan_payload = None
        if p:
            plan_payload = {
                'id': p.id,
                'name': p.name,
                'price': float(p.price),
                'max_systems': p.max_systems,
                'max_tables_per_system': p.max_tables_per_system,
                'max_storage_mb': p.max_storage_mb,
                'has_ai_assistant': bool(p.has_ai_assistant),
            }

        return Response({
            'id': user.id, 'name': user.name, 'email': user.email,
            'avatarUrl': user.avatar_url or '',
            'phone': user.phone or '',
            'planId': user.plan_id, 'planName': plan_name,
            'plan': plan_payload,
            'role': getattr(user, 'role', 'user'),
            'expertise_level': getattr(user, 'expertise_level', 'beginner'),
            'createdAt': user.created_at.isoformat() if user.created_at else None,
            'needsTermsAcceptance': needs_terms_acceptance,
            'termsContent': terms_content,
            'termsVersion': current_terms_version,
            'hasAi': user_has_ai(user),
        })
    else:
        name = request.data.get('name')
        phone = request.data.get('phone')
        if name is not None:
            err = val.validate_profile_name(name)
            if err:
                return Response({'error': err}, status=400)
            user.name = name
        if phone is not None:
            err = val.validate_profile_phone(phone)
            if err:
                return Response({'error': err}, status=400)
            user.phone = phone if str(phone).strip() else None
        expertise = request.data.get('expertise_level')
        if expertise is not None and expertise in ('beginner', 'intermediate', 'advanced'):
            user.expertise_level = expertise
        user.save()
        return Response({'ok': True})


@api_view(['PUT'])
def user_expertise_view(request):
    user, err = require_auth(request)
    if err:
        return err
    level = request.data.get('level', '').strip().lower()
    if level not in ('beginner', 'intermediate', 'advanced'):
        return Response({'error': 'Nivel invalido. Usa: beginner, intermediate, advanced'}, status=400)
    user.expertise_level = level
    user.save()
    return Response({'ok': True, 'expertise_level': level})


@api_view(['PUT'])
def user_password_view(request):
    user, err = require_auth(request)
    if err:
        return err
    if is_password_verify_locked(user.id):
        return Response(
            {'error': 'Demasiados intentos. Espera unos minutos antes de cambiar la contraseña.'},
            status=status.HTTP_429_TOO_MANY_REQUESTS,
        )
    current = request.data.get('currentPassword')
    new_pwd = request.data.get('newPassword')
    if not _verify_password(current, user.password_hash):
        record_password_verify_failure(user.id)
        return Response({'error': 'Password actual incorrecta'}, status=400)
    clear_password_verify_failures(user.id)
    err = val.validate_new_password(new_pwd or '')
    if err:
        return Response({'error': err}, status=400)
    user.password_hash = make_password(new_pwd)
    user.save()
    return Response({'ok': True})


@api_view(['PUT'])
def user_avatar_view(request):
    user, err = require_auth(request)
    if err:
        return err
    user.avatar_url = request.data.get('avatarUrl')
    user.save()
    return Response({'ok': True})


@api_view(['PUT'])
def user_plan_view(request):
    user, err = require_auth(request)
    if err:
        return err
    sync_plans_into_db()
    plan_id = request.data.get('newPlanId')
    plan = Plan.objects.filter(id=plan_id, is_active=True).first()
    if not plan and isinstance(plan_id, str):
        plan = Plan.objects.filter(name__iexact=plan_id, is_active=True).first()

    if not plan or plan.name not in ('Free', 'Pro', 'Corporate'):
        return Response({'error': 'Plan no disponible'}, status=404)
    user.plan = plan
    user.save()
    return Response({
        'ok': True,
        'planName': plan.name,
        'planId': plan.id,
        'hasAi': user_has_ai(user),
        'plan': {
            'id': plan.id,
            'name': plan.name,
            'price': float(plan.price),
            'max_systems': plan.max_systems,
            'max_tables_per_system': plan.max_tables_per_system,
            'max_storage_mb': plan.max_storage_mb,
            'has_ai_assistant': bool(plan.has_ai_assistant),
        },
    })


@api_view(['POST'])
def user_verify_password_view(request):
    user, err = require_auth(request)
    if err:
        return err
    if is_password_verify_locked(user.id):
        return Response(
            {'error': 'Demasiados intentos. Espera unos minutos.'},
            status=status.HTTP_429_TOO_MANY_REQUESTS,
        )
    password = request.data.get('password')
    if _verify_password(password, user.password_hash):
        clear_password_verify_failures(user.id)
        return Response({'ok': True})
    record_password_verify_failure(user.id)
    return Response({'error': 'Contraseña incorrecta'}, status=401)


# ═══════════════════════════════════════════
# SYSTEMS
# ═══════════════════════════════════════════

@api_view(['GET', 'POST'])
def systems_list_view(request):
    user, err = require_auth(request)
    if err: return err

    if request.method == 'GET':
        # Sistemas propios + sistemas donde soy colaborador
        owned = System.objects.filter(owner=user).select_related('owner')
        collab_ids = SystemCollaborator.objects.filter(user=user, can_read=True).values_list('system_id', flat=True)
        collab_systems = System.objects.filter(id__in=collab_ids).select_related('owner')
        
        all_systems = (owned | collab_systems).distinct().order_by('-created_at')
        return Response([serialize_system(s, user) for s in all_systems])
    else:
        name = (request.data.get('name') or '').strip()
        if not name:
            return Response({'error': 'Nombre requerido'}, status=status.HTTP_400_BAD_REQUEST)
        errn = val.validate_system_name(name)
        if errn:
            return Response({'error': errn}, status=status.HTTP_400_BAD_REQUEST)
        sec_mode = request.data.get('securityMode', 'none') or 'none'
        if sec_mode == 'general':
            gp = request.data.get('generalPassword') or ''
            errp = val.validate_system_general_password(gp)
            if errp:
                return Response({'error': errp}, status=status.HTTP_400_BAD_REQUEST)
        try:
            assert_can_create_system(user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_403_FORBIDDEN)
        idem = (request.headers.get('X-Idempotency-Key') or request.data.get('idempotencyKey') or '').strip()
        if idem:
            sk = f'sys_idem_{user.id}_{idem}'
            hit = request.session.get(sk)
            if isinstance(hit, dict) and (time.time() - hit.get('t', 0)) < 3600:
                sid = hit.get('id')
                if sid:
                    ex = System.objects.filter(id=sid, owner=user, is_deleted=False).first()
                    if ex:
                        return Response(serialize_system(ex), status=status.HTTP_200_OK)
        dup = System.objects.filter(owner=user, name=name, is_deleted=False).first()
        if dup:
            return Response(
                {'error': 'Ya existe un sistema con este nombre', 'existing': serialize_system(dup)},
                status=status.HTTP_409_CONFLICT,
            )
        try:
            with transaction.atomic():
                s = System.objects.create(
                    owner=user,
                    name=name,
                    description=request.data.get('description'),
                    image_url=request.data.get('imageUrl'),
                    security_mode=request.data.get('securityMode', 'none'),
                    general_password=request.data.get('generalPassword'),
                )
        except IntegrityError:
            dup = System.objects.filter(owner=user, name=name, is_deleted=False).first()
            if dup:
                return Response(serialize_system(dup), status=status.HTTP_200_OK)
            return Response({'error': 'No se pudo crear el sistema'}, status=status.HTTP_409_CONFLICT)
        log_action(user, s, 'CREAR_SISTEMA', f'Sistema "{name}" creado')
        if idem:
            request.session[f'sys_idem_{user.id}_{idem}'] = {'id': s.id, 't': time.time()}
            request.session.modified = True
        return Response(serialize_system(s), status=status.HTTP_201_CREATED)


@api_view(['POST'])
def table_records_bulk_delete_view(request, table_id):
    user = get_current_user(request)
    if not user: return Response({'error': 'No autorizado'}, status=401)
    
    table = get_object_or_404(SystemTable, id=table_id)
    # Re-check permission logic helper
    has_perm, err_res = check_table_permission(user, table, 'delete')
    if not has_perm:
        return err_res

    record_ids = request.data.get('ids', [])
    if not record_ids:
        return Response({'error': 'No se proporcionaron IDs'}, status=400)

    try:
        with transaction.atomic():
            count, _ = SystemRecord.objects.filter(table=table, id__in=record_ids).delete()
            log_action(user, table.system, 'ELIMINAR_MASIVO', f'Eliminados {count} registros de la tabla "{table.name}"')
            return Response({'ok': True, 'deletedCount': count})
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['GET', 'PUT', 'DELETE'])
def systems_detail_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    
    try:
        s = System.objects.get(id=pk)
    except System.DoesNotExist:
        return Response({'error': 'Sistema no encontrado'}, status=404)

    if request.method == 'GET':
        ok, res = check_system_permission(user, pk, 'read')
        if not ok: return res
        return Response(serialize_system(s, user))
    elif request.method == 'PUT':
        if s.owner_id != user.id:
            return Response({'error': 'Solo el propietario puede editar el sistema.'}, status=403)
        name = (request.data.get('name') or s.name or '').strip()
        errn = val.validate_system_name(name)
        if errn:
            return Response({'error': errn}, status=400)
        sec_mode = request.data.get('securityMode', s.security_mode) or 'none'
        if sec_mode == 'general':
            gp = request.data.get('generalPassword') or s.general_password or ''
            errp = val.validate_system_general_password(gp)
            if errp:
                return Response({'error': errp}, status=400)
        s.name = name
        s.description = request.data.get('description', s.description)
        s.image_url = request.data.get('imageUrl', s.image_url)
        s.security_mode = sec_mode
        pwd = request.data.get('generalPassword')
        if pwd:
            s.general_password = pwd
        s.save()
        log_action(user, s, 'EDITAR_SISTEMA', f'Sistema "{s.name}" editado')
        return Response(serialize_system(s, user))
    else:
        if s.owner_id != user.id:
            return Response({'error': 'Solo el propietario puede eliminar el sistema.'}, status=403)
        name = s.name
        log_action(user, s, 'ELIMINAR_SISTEMA', f'Sistema "{name}" eliminado')
        s.delete()
        return Response({'ok': True})


@api_view(['GET'])
def systems_estadisticas_view(request):
    user, err = require_auth(request)
    if err:
        return err

    systems = System.objects.filter(owner=user)
    total_systems = systems.count()
    tables = SystemTable.objects.filter(system__in=systems)
    total_records = SystemRecord.objects.filter(table__in=tables).count()

    sec_none = systems.filter(Q(security_mode='none') | Q(security_mode__isnull=True)).count()
    sec_general = systems.filter(security_mode='general').count()
    sec_individual = systems.filter(security_mode='individual').count()

    plan = plan_for_user(user)
    plan_usage = {
        'planName': plan.name if plan else 'Gratuito',
        'current': total_systems,
        'max': plan.max_systems if plan else 1,
    }

    today = timezone.now().date()
    labels = []
    data = []
    day_names = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        labels.append(day_names[d.weekday()])
        count = AuditLog.objects.filter(system__in=systems, created_at__date=d).count()
        data.append(count)

    return Response({
        'totalSystems': total_systems, 'totalUsers': 1,
        'totalRecords': total_records,
        'securityNone': sec_none, 'securityGeneral': sec_general,
        'securityIndividual': sec_individual,
        'planUsage': plan_usage,
        'activityLabels': labels, 'activityData': data,
    })


@api_view(['POST'])
def system_verify_password_view(request, pk):
    user, err = require_auth(request)
    if err:
        return err
    try:
        s = System.objects.get(id=pk)
    except System.DoesNotExist:
        return Response({'error': 'Sistema no encontrado'}, status=404)
    password = request.data.get('password')
    if s.general_password and s.general_password == password:
        return Response({'ok': True})
    return Response({'error': 'Contraseña incorrecta'}, status=401)


# ═══════════════════════════════════════════
# SYSTEM TABLES
# ═══════════════════════════════════════════

@api_view(['GET', 'POST'])
def system_tables_view(request, pk):
    user, err = require_auth(request)
    if err: return err

    if request.method == 'GET':
        ok, res = check_system_permission(user, pk, 'read')
        if not ok: return res
        tables = tables_queryset_for_user(user, pk)
        return Response([serialize_table(t, user) for t in tables])
    else:
        try:
            sys_obj = System.objects.get(id=pk)
        except System.DoesNotExist:
            return Response({'error': 'Sistema no encontrado'}, status=404)
        if sys_obj.owner_id != user.id:
            return Response({'error': 'Solo el propietario puede crear tablas.'}, status=403)
        ok, res = check_system_permission(user, pk, 'read')
        if not ok: return res
        try:
            assert_can_add_table(sys_obj)
        except ValueError as e:
            return Response({'error': str(e)}, status=403)
        name = (request.data.get('name') or '').strip()
        errn = val.validate_table_name(name)
        if errn:
            return Response({'error': errn}, status=400)
        desc = request.data.get('description')
        fields_data = request.data.get('fields', [])
        table = SystemTable.objects.create(system_id=pk, name=name, description=desc)
        _save_fields(table, fields_data)
        log_action(user, table.system, 'CREAR_TABLA', f'Tabla "{name}" creada')
        return Response(serialize_table(table, user), status=201)


@api_view(['PUT', 'DELETE'])
def system_table_detail_view(request, system_pk, table_pk):
    user, err = require_auth(request)
    if err: return err

    try:
        table = SystemTable.objects.get(id=table_pk, system_id=system_pk)
    except SystemTable.DoesNotExist:
        return Response({'error': 'Tabla no encontrada'}, status=404)

    if request.method == 'PUT':
        if table.system.owner_id != user.id:
            return Response({'error': 'Solo el propietario puede editar tablas.'}, status=403)
        ok, res = check_system_permission(user, system_pk, 'read')
        if not ok: return res
        table.name = request.data.get('name', table.name)
        table.description = request.data.get('description', table.description)
        table.save()
        fields_data = request.data.get('fields', [])
        if fields_data:
            _save_fields(table, fields_data, update=True)
        log_action(user, table.system, 'EDITAR_TABLA', f'Tabla "{table.name}" editada')
        return Response(serialize_table(table, user))
    else:
        if table.system.owner_id != user.id:
            return Response({'error': 'Solo el propietario puede eliminar tablas.'}, status=403)
        ok, res = check_system_permission(user, system_pk, 'read')
        if not ok: return res
        name = table.name
        system = table.system
        table.delete()
        log_action(user, system, 'ELIMINAR_TABLA', f'Tabla "{name}" eliminada')
        return Response({'ok': True})


def _save_fields(table, fields_data, update=False):
    if update:
        existing_ids = {fd.get('id') for fd in fields_data if fd.get('id')}
        SystemField.objects.filter(table=table).exclude(id__in=existing_ids).delete()

    # Integrity Check: Only one PK per table
    pk_count = sum(1 for fd in fields_data if fd.get('isPrimaryKey'))
    if pk_count > 1:
        raise exceptions.ValidationError("Una tabla solo puede tener una llave primaria.")

    for i, fd in enumerate(fields_data):
        fid = fd.get('id')
        name = fd.get('name', '')
        ftype = fd.get('type', 'text')
        required = fd.get('required', False)
        options = fd.get('options', [])
        related_table_id = fd.get('relatedTableId')
        related_display_field_id = fd.get('relatedDisplayFieldId')
        is_pk = fd.get('isPrimaryKey', False)

        # Integrity Check: Relation targets must have a PK
        if ftype == 'relation' and related_table_id:
            # Check if target table has a PK
            target_pk = SystemField.objects.filter(table_id=related_table_id, is_primary_key=True).first()
            # If creating a new table and relating to it (edge case) or if it's an existing table
            if not target_pk:
                # We also need to check if the PK is being defined in the same transaction for the SAME table 
                # (but relation typically points to DIFFERENT tables).
                # If target is DIFFERENT table, we check its current state.
                if int(related_table_id) != table.id:
                    raise exceptions.ValidationError(f"La tabla de destino de la relación '{name}' debe tener una Llave Primaria (PK) definida.")

        defaults = dict(
            name=name, type=ftype, required=required, 
            is_unique=fd.get('unique', False) or is_pk,
            is_primary_key=is_pk,
            order_index=i,
            related_table_id=related_table_id,
            related_display_field_id=related_display_field_id,
        )

        if fid and update:
            field, _ = SystemField.objects.update_or_create(
                id=fid, table=table, defaults=defaults
            )
        else:
            field = SystemField.objects.create(table=table, **defaults)

        SystemFieldOption.objects.filter(field=field).delete()
        if ftype == 'select' and options:
            for opt in options:
                SystemFieldOption.objects.create(field=field, value=opt)


# ═══════════════════════════════════════════
# SYSTEM INVITATIONS (stubs)
# ═══════════════════════════════════════════

@api_view(['GET'])
def system_invitations_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    try:
        sys = System.objects.get(id=pk)
    except System.DoesNotExist:
        return Response({'error': 'Sistema no encontrado'}, status=404)
        
    # Solo el dueño puede ver la lista de invitados completa
    if sys.owner_id != user.id:
        return Response({'error': 'No autorizado'}, status=403)
        
    collabs = SystemCollaborator.objects.filter(system=sys).select_related('user')
    data = []
    for c in collabs:
        trows = SystemCollaboratorTable.objects.filter(collaborator=c).select_related('table')
        table_perms = [
            {
                'tableId': r.table_id,
                'tableName': r.table.name,
                'can_read': r.can_read,
                'can_create': r.can_create,
                'can_update': r.can_update,
                'can_delete': r.can_delete,
            }
            for r in trows
        ]
        data.append({
            'id': c.id,
            'email': c.user.email,
            'name': c.user.name,
            'can_read': c.can_read,
            'can_create': c.can_create,
            'can_update': False,
            'can_delete': False,
            'tablePermissions': table_perms,
        })
    return Response(data)


@api_view(['POST'])
def system_invite_view(request, pk):
    user, err = require_auth(request)
    if err: return err

    try:
        sys = System.objects.get(id=pk)
    except System.DoesNotExist:
        return Response({'error': 'Sistema no encontrado'}, status=404)

    if sys.owner_id != user.id:
        return Response({'error': 'No autorizado para invitar'}, status=403)

    owner_pw = request.data.get('ownerPassword') or request.data.get('password')
    if not owner_pw or not _verify_password(owner_pw, user.password_hash):
        return Response({'error': 'Contraseña del propietario incorrecta'}, status=403)

    email = (request.data.get('email') or '').strip()
    if not email:
        return Response({'error': 'Email es requerido'}, status=400)

    try:
        target_user = User.objects.get(email__iexact=email)
    except User.DoesNotExist:
        return Response({'error': 'Usuario no encontrado en Datium'}, status=404)

    if target_user.id == user.id:
        return Response({'error': 'No puedes invitarte a ti mismo'}, status=400)

    can_read = request.data.get('can_read', True)
    can_create = request.data.get('can_create', False)
    can_update = False
    can_delete = False

    collab, created = SystemCollaborator.objects.get_or_create(
        system=sys, user=target_user,
        defaults={
            'can_read': can_read,
            'can_create': can_create,
            'can_update': can_update,
            'can_delete': can_delete,
        },
    )

    if not created:
        collab.can_read = can_read
        collab.can_create = can_create
        collab.can_update = can_update
        collab.can_delete = can_delete
        collab.save()

    table_perms = request.data.get('tablePerms') or []
    if isinstance(table_perms, list) and table_perms:
        SystemCollaboratorTable.objects.filter(collaborator=collab).delete()
        for row in table_perms:
            tid = row.get('tableId')
            if not tid:
                continue
            if not SystemTable.objects.filter(id=tid, system=sys).exists():
                continue
            SystemCollaboratorTable.objects.create(
                collaborator=collab,
                table_id=tid,
                can_read=bool(row.get('read', row.get('can_read', True))),
                can_create=bool(row.get('create', row.get('can_create', False))),
                can_update=bool(row.get('update', row.get('can_update', False))),
                can_delete=bool(row.get('delete', row.get('can_delete', False))),
            )

    return Response({'ok': True, 'message': 'Colaborador actualizado/añadido'}, status=201 if created else 200)


@api_view(['DELETE'])
def system_invitation_delete_view(request, system_pk, share_pk):
    user, err = require_auth(request)
    if err: return err
    
    try:
        collab = SystemCollaborator.objects.get(id=share_pk, system_id=system_pk)
    except SystemCollaborator.DoesNotExist:
        return Response({'error': 'Colaboración no encontrada'}, status=404)
        
    # El dueño del sistema o el mismo colaborador pueden borrarla
    if collab.system.owner_id != user.id and collab.user_id != user.id:
        return Response({'error': 'No autorizado'}, status=403)
        
    collab.delete()
    return Response({'ok': True})


# ═══════════════════════════════════════════
# TABLE ENDPOINTS
# ═══════════════════════════════════════════

@api_view(['GET'])
def table_detail_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    try:
        table = SystemTable.objects.get(id=pk)
    except SystemTable.DoesNotExist:
        return Response({'error': 'Tabla no encontrada'}, status=404)
        
    ok, res = check_table_permission(user, table, 'read')
    if not ok: return res

    return Response(serialize_table(table, user))


@api_view(['GET'])
def table_fields_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    
    try:
        table = SystemTable.objects.get(id=pk)
    except SystemTable.DoesNotExist:
        return Response({'error': 'Tabla no encontrada'}, status=404)
        
    ok, res = check_table_permission(user, table, 'read')
    if not ok: return res

    fields = SystemField.objects.filter(table_id=pk).order_by('order_index')
    return Response([serialize_field(f) for f in fields])


@api_view(['GET', 'POST'])
def table_records_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    try:
        table = SystemTable.objects.get(id=pk)
    except SystemTable.DoesNotExist:
        return Response({'error': 'Tabla no encontrada'}, status=404)

    fields = list(SystemField.objects.filter(table=table).order_by('order_index'))

    if request.method == 'GET':
        ok, res = check_table_permission(user, table, 'read')
        if not ok: return res
        records = SystemRecord.objects.filter(table=table).order_by('-created_at')
        return Response([serialize_record(r, fields) for r in records])
    else:
        ok, res = check_table_permission(user, table, 'create')
        if not ok: return res
        values = request.data.get('values', {})
        
        err_msg = _validate_record_values(table, fields, values)
        if err_msg:
            return Response({'error': err_msg}, status=400)
            
        record = SystemRecord.objects.create(table=table, created_by=user)
        for field in fields:
            val = values.get(str(field.id))
            if val is None:
                val = values.get(field.name)
            
            if val is not None:
                SystemRecordValue.objects.create(record=record, field=field, value=str(val))
                
        log_action(user, table.system, 'CREAR_REGISTRO', f'Registro #{record.id} en tabla "{table.name}"')
        return Response(serialize_record(record, fields), status=201)


@api_view(['PUT', 'DELETE'])
def table_record_detail_view(request, table_pk, record_pk):
    user, err = require_auth(request)
    if err: return err
    try:
        record = SystemRecord.objects.get(id=record_pk, table_id=table_pk)
    except SystemRecord.DoesNotExist:
        return Response({'error': 'Registro no encontrado'}, status=404)

    table = record.table
    fields = list(SystemField.objects.filter(table=table).order_by('order_index'))

    if request.method == 'PUT':
        ok, res = check_table_permission(user, table, 'update')
        if not ok: return res
        values = request.data.get('values', {})
        
        err_msg = _validate_record_values(table, fields, values, record_id=record.id)
        if err_msg:
            return Response({'error': err_msg}, status=400)
            
        for field in fields:
            val = values.get(str(field.id))
            if val is None:
                val = values.get(field.name)
            
            if val is not None:
                SystemRecordValue.objects.update_or_create(
                    record=record, field=field, defaults={'value': str(val)}
                )
                
        log_action(user, table.system, 'EDITAR_REGISTRO', f'Registro #{record.id} editado')
        return Response(serialize_record(record, fields))
    else:
        ok, res = check_table_permission(user, table, 'delete')
        if not ok: return res
        rid = record.id
        record.delete()
        log_action(user, table.system, 'ELIMINAR_REGISTRO', f'Registro #{rid} eliminado')
        return Response({'ok': True})


@api_view(['GET'])
def table_export_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    try:
        table = SystemTable.objects.get(id=pk)
    except SystemTable.DoesNotExist:
        return Response({'error': 'Tabla no encontrada'}, status=404)

    ok, res = check_table_permission(user, table, 'read')
    if not ok: return res

    fmt = request.GET.get('format', 'csv').lower()
    fields = list(SystemField.objects.filter(table=table).order_by('order_index'))
    # Use prefetch_related to solve N+1 Problem
    records = SystemRecord.objects.filter(table=table).prefetch_related('values').order_by('created_at')

    filename = f"datium_export_{table.name.lower().replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}"

    if fmt == 'json':
        data = [serialize_record(r, fields) for r in records]
        response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        return response

    elif fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = table.name[:30] # Excel limit
        
        # Headers
        headers = ['ID'] + [f.name for f in fields]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='137fec', end_color='137fec', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data Rows
        for record in records:
            vals = {v.field_id: v.value for v in record.values.all()}
            row_data = [record.id] + [vals.get(f.id, '') for f in fields]
            ws.append(row_data)
        
        # Adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap at 50

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    elif fmt == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Datium System: {table.system.name}", styles['Title']))
        elements.append(Paragraph(f"Tabla: {table.name}", styles['Heading2']))
        elements.append(Paragraph(f"Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Data preparation
        data = [['ID'] + [f.name for f in fields]]
        for record in records:
            vals = {v.field_id: v.value for v in record.values.all()}
            row = [str(record.id)]
            for f in fields:
                val = str(vals.get(f.id, ''))
                row.append(val[:30] if len(val) > 30 else val)
            data.append(row)
        
        # Create Table
        t = Table(data, hAlign='LEFT')
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#137fec')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(t)
        doc.build(elements)
        return response

    else: # Default CSV
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID'] + [f.name for f in fields])
        for record in records:
            vals = {v.field_id: v.value for v in record.values.all()}
            writer.writerow([record.id] + [vals.get(f.id, '') for f in fields])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response


@api_view(['PUT'])
def table_move_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    target_system_id = request.GET.get('targetSystemId')
    if not target_system_id:
        return Response({'error': 'Sistema destino requerido'}, status=400)
    try:
        source_table = SystemTable.objects.get(id=pk)
        target_system = System.objects.get(id=target_system_id)
    except (SystemTable.DoesNotExist, System.DoesNotExist):
        return Response({'error': 'Tabla o sistema no encontrado'}, status=404)

    ok, res = check_table_permission(user, source_table, 'read')
    if not ok: return res
    if target_system.owner_id != user.id:
        return Response({'error': 'Solo el propietario del sistema destino puede recibir tablas.'}, status=403)
    try:
        assert_can_add_table(target_system)
    except ValueError as e:
        return Response({'error': str(e)}, status=403)

    new_table = SystemTable.objects.create(
        system=target_system, name=source_table.name, description=source_table.description
    )
    field_map = {}
    for field in SystemField.objects.filter(table=source_table).order_by('order_index'):
        nf = SystemField.objects.create(
            table=new_table, name=field.name, type=field.type,
            required=field.required, order_index=field.order_index,
            related_table_id=field.related_table_id,
            related_display_field_id=field.related_display_field_id,
        )
        field_map[field.id] = nf.id
        for opt in SystemFieldOption.objects.filter(field=field):
            SystemFieldOption.objects.create(field=nf, value=opt.value)

    for record in SystemRecord.objects.filter(table=source_table):
        nr = SystemRecord.objects.create(table=new_table, created_by=user)
        for rv in SystemRecordValue.objects.filter(record=record):
            nfid = field_map.get(rv.field_id)
            if nfid:
                SystemRecordValue.objects.create(record=nr, field_id=nfid, value=rv.value)

    log_action(user, target_system, 'COPIAR_TABLA', f'Tabla "{source_table.name}" copiada')
    return Response(serialize_table(new_table, user))


@api_view(['POST'])
def table_bulk_import_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    try:
        table = SystemTable.objects.get(id=pk)
    except SystemTable.DoesNotExist:
        return Response({'error': 'Tabla no encontrada'}, status=404)

    ok, res = check_table_permission(user, table, 'create')
    if not ok: return res

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No se proporcionó ningún archivo'}, status=400)

    fields = list(SystemField.objects.filter(table=table).order_by('order_index'))
    field_name_map = {f.name.lower(): f for f in fields}
    
    filename = file.name.lower()
    records_to_create_data = []
    
    try:
        if filename.endswith('.csv'):
            decoded_file = file.read().decode('utf-8')
            reader = csv.DictReader(StringIO(decoded_file))
            records_to_create_data = list(reader)
        elif filename.endswith('.json'):
            records_to_create_data = json.load(file)
            if not isinstance(records_to_create_data, list):
                return Response({'error': 'JSON debe ser una lista de objetos'}, status=400)
        else:
            return Response({'error': 'Formato de archivo no soportado. Use CSV o JSON'}, status=400)

        created_count = 0
        with transaction.atomic():
            for rec_data in records_to_create_data:
                if not isinstance(rec_data, dict): continue
                data_map = rec_data.get('fieldValues', rec_data)
                if not isinstance(data_map, dict): continue
                
                record = SystemRecord.objects.create(table=table, created_by=user)
                
                # Batch create values for this record
                values_to_create = []
                for key, val in data_map.items():
                    field = field_name_map.get(key.lower())
                    if field:
                        values_to_create.append(SystemRecordValue(
                            record=record, 
                            field=field, 
                            value=str(val) if val is not None else ''
                        ))
                
                if values_to_create:
                    SystemRecordValue.objects.bulk_create(values_to_create)
                
                created_count += 1
        
        log_action(user, table.system, 'IMPORTAR_DATOS', f'Importados {created_count} registros en "{table.name}"')
        return Response({'ok': True, 'count': created_count})
    except Exception as e:
        return Response({'error': f'Error al procesar archivo: {str(e)}'}, status=500)


# ═══════════════════════════════════════════
# AUDIT
# ═══════════════════════════════════════════

def _filter_logs(request, system_ids):
    search = request.GET.get('search', '')
    date_from = request.GET.get('dateFrom', '')
    date_to = request.GET.get('dateTo', '')
    user_id = request.GET.get('userId', '')

    # Auto-cleanup logs older than 60 days
    limit_date = timezone.now() - timedelta(days=60)
    AuditLog.objects.filter(created_at__lt=limit_date).delete()
    SecurityAudit.objects.filter(created_at__lt=limit_date).delete()

    qs = AuditLog.objects.filter(system_id__in=system_ids).order_by('-created_at')
    if search:
        qs = qs.filter(Q(action__icontains=search) | Q(details__icontains=search))
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if user_id:
        qs = qs.filter(user_id=user_id)

    return [{
        'id': l.id, 'action': l.action, 'details': l.details or '',
        'ip': l.ip or '', 'createdAt': l.created_at.isoformat(),
        'userName': l.user.name if l.user else 'Desconocido',
        'systemName': l.system.name if l.system else '',
    } for l in qs[:100]]


def _filter_security(request, system_ids):
    search = request.GET.get('search', '')
    date_from = request.GET.get('dateFrom', '')
    date_to = request.GET.get('dateTo', '')
    user_id = request.GET.get('userId', '')

    qs = SecurityAudit.objects.filter(system_id__in=system_ids).order_by('-created_at')
    if search:
        qs = qs.filter(Q(event__icontains=search) | Q(details__icontains=search))
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if user_id:
        qs = qs.filter(user_id=user_id)

    return [{
        'id': l.id, 'event': l.event, 'details': l.details or '',
        'severity': l.severity, 'createdAt': l.created_at.isoformat(),
        'userName': l.user.name if l.user else 'Desconocido',
        'systemName': l.system.name if l.system else '',
    } for l in qs[:100]]


@api_view(['GET'])
def audit_logs_global_view(request):
    user, err = require_auth(request)
    if err:
        return err
    system_ids = System.objects.filter(owner=user).values_list('id', flat=True)
    return Response(_filter_logs(request, system_ids))


@api_view(['GET'])
def audit_security_global_view(request):
    user, err = require_auth(request)
    if err:
        return err
    system_ids = System.objects.filter(owner=user).values_list('id', flat=True)
    return Response(_filter_security(request, system_ids))


@api_view(['GET'])
def audit_logs_system_view(request, pk):
    user, err = require_auth(request)
    if err:
        return err
    return Response(_filter_logs(request, [pk]))


@api_view(['GET'])
def audit_security_system_view(request, pk):
    user, err = require_auth(request)
    if err:
        return err
    return Response(_filter_security(request, [pk]))


@api_view(['GET'])
def audit_system_users_view(request, pk):
    user, err = require_auth(request)
    if err:
        return err
    user_ids = AuditLog.objects.filter(system_id=pk).values_list('user_id', flat=True).distinct()
    users = User.objects.filter(id__in=user_ids)
    return Response([{'userId': u.id, 'name': u.name or 'Sin nombre', 'email': u.email} for u in users])


# ═══════════════════════════════════════════
# IMAGE UPLOAD
# ═══════════════════════════════════════════

@api_view(['POST'])
def upload_image_view(request):
    user, err = require_auth(request)
    if err:
        return err
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No se proporcionó archivo'}, status=400)

    err_img = val.validate_image_extension(file.name)
    if err_img:
        return Response({'error': err_img}, status=400)

    plan = plan_for_user(user)
    limit_bytes = plan.max_storage_mb * 1024 * 1024 if plan else 1024 * 1024 * 1024
    if user.storage_used_bytes + file.size > limit_bytes:
        return Response({'error': 'Has excedido el límite de almacenamiento de tu plan.'}, status=403)

    user.storage_used_bytes += file.size
    user.save()

    upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
    os.makedirs(upload_dir, exist_ok=True)

    ext = os.path.splitext(file.name)[1]
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(upload_dir, filename)

    with open(filepath, 'wb+') as dest:
        for chunk in file.chunks():
            dest.write(chunk)

    url = f"{settings.MEDIA_URL}uploads/{filename}"
    return Response({'url': url})


# ═══════════════════════════════════════════
# REPORTS & ADMIN
# ═══════════════════════════════════════════




@api_view(['POST'])
def user_reports_view(request):
    user, err = require_auth(request)
    if err: return err
    title = (request.data.get('title') or '').strip()
    summary = (request.data.get('summary') or '').strip()
    screenshot_url = request.data.get('screenshot_url', '')
    err_rep = val.validate_report_fields(title, summary)
    if err_rep:
        return Response({'error': err_rep}, status=400)
    rep = UserReport.objects.create(
        user=user, title=title, summary=summary, screenshot_url=screenshot_url
    )
    return Response({'ok': True, 'reportId': rep.id}, status=201)


@api_view(['PUT'])
def admin_report_detail_view(request, pk):
    user, err = require_admin(request)
    if err: return err
    try:
        rep = UserReport.objects.get(id=pk)
    except UserReport.DoesNotExist:
        return Response({'error': 'Reporte no encontrado'}, status=404)
    rep.status = request.data.get('status', rep.status)
    rep.save()
    return Response({'ok': True, 'status': rep.status})


@api_view(['GET'])
def admin_stats_view(request):
    user, err = require_admin(request)
    if err:
        return err
    return Response({
        'users': User.objects.count(),
        'systems': System.objects.filter(is_deleted=False).count(),
        'reports': UserReport.objects.filter(status='pending').count(),
    })


@api_view(['GET', 'POST'])
def admin_reports_view(request):
    user, err = require_admin(request)
    if err: return err
    if request.method == 'POST':
        rid = request.data.get('report_id')
        new_status = request.data.get('status')
        if rid and new_status:
            UserReport.objects.filter(id=rid).update(status=new_status)
        return JsonResponse({'status': 'success'})
    reports = UserReport.objects.select_related('user').all().order_by('-created_at')
    data = [{
        'id': r.id,
        'user': r.user.email if r.user else '',
        'userEmail': r.user.email if r.user else '',
        'title': r.title,
        'summary': r.summary,
        'screenshot_url': r.screenshot_url,
        'status': r.status,
        'createdAt': r.created_at.isoformat(),
    } for r in reports]
    return Response(data)

@api_view(['POST'])
def import_records_view(request, table_id: int):
    """Mass import records from JSON/CSV using bulk_create."""
    user, err = require_auth(request)
    if err: return err
    
    table = SystemTable.objects.filter(id=table_id).first()
    if not table: return Response({'error': 'Tabla no encontrada'}, status=404)
    
    ok, res = check_table_permission(user, table, 'create')
    if not ok: return res

    records_data = request.data.get('records', [])
    if not isinstance(records_data, list):
        return Response({'error': 'Formato de datos inválido. Se espera una lista.'}, status=400)

    fields_map = {f.name.lower(): f for f in SystemField.objects.filter(table=table)}
    records_to_create = []
    values_to_create = []
    ip = request.META.get("REMOTE_ADDR", "")

    try:
        with transaction.atomic():
            for item in records_data:
                rec = SystemRecord(table=table, created_by=user)
                records_to_create.append(rec)
            created_records = SystemRecord.objects.bulk_create(records_to_create)
            
            for i, item in enumerate(records_data):
                rec = created_records[i]
                for k, v in item.items():
                    f = fields_map.get(k.lower())
                    if f and v is not None and str(v).strip() != "":
                        values_to_create.append(SystemRecordValue(record=rec, field=f, value=str(v)))
            
            # Second bulk create values
            SystemRecordValue.objects.bulk_create(values_to_create)
            
            AuditLog.objects.create(
                user=user, system=table.system, action="IMPORT_RECORDS",
                details=f"Importación masiva: {len(created_records)} registros en {table.name}",
                ip=ip
            )
        return Response({'status': 'success', 'imported': len(created_records)})
    except Exception as e:
        return Response({'error': f"Error durante la importación: {str(e)}"}, status=500)

def generate_export_response(header, rows, filename_base, format_param, title="Datium Export"):
    """Universal helper to generate export responses for CSV, XLSX, and PDF."""
    now_str = timezone.now().strftime("%Y%m%d_%H%M")
    
    if format_param == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}_{now_str}.csv"'
        writer = csv.writer(response)
        writer.writerow(header)
        writer.writerows(rows)
        return response

    if format_param == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        safe_title = "".join([c for c in title if c not in r'\/*?:[]'])
        ws.title = safe_title[:31]
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws.column_dimensions[column].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}_{now_str}.xlsx"'
        wb.save(response)
        return response

    if format_param == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
        elements.append(Paragraph(title, title_style))
        table_data = [header] + rows
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
        ]))
        elements.append(t)
        doc.build(elements)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}_{now_str}.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    return JsonResponse({'status': 'error', 'message': 'Formato no soportado'}, status=400)

def export_records_view(request, table_id: int):
    """Export records from table."""
    user, err = require_auth(request)
    if err: 
        return JsonResponse({'error': 'No autenticado'}, status=401)
    
    table = SystemTable.objects.filter(id=table_id).first()
    if not table: return JsonResponse({'error': 'Tabla no encontrada (ID Inválido o Eliminada)'}, status=404)

    ok, res = check_table_permission(user, table, 'read')
    if not ok: return JsonResponse({'error': 'Accesos insuficientes en la tabla'}, status=403)

    records = SystemRecord.objects.filter(table=table).prefetch_related('systemrecordvalue_set__field')
    fields = SystemField.objects.filter(table=table).order_by('order_index')
    
    format_param = request.GET.get('format', 'json').lower()
    
    header = ['ID'] + [f.name for f in fields]
    rows = []
    for r in records:
        vals = {v.field.name: v.value for v in r.systemrecordvalue_set.all()}
        row = [r.id] + [vals.get(f.name, "") for f in fields]
        rows.append(row)

    if format_param in ['csv', 'xlsx', 'pdf']:
        return generate_export_response(header, rows, f"export_{table.name}", format_param, f"Tabla: {table.name}")

    data = []
    for i, r in enumerate(records):
        row_data = {'id': r.id}
        for j, val in enumerate(rows[i][1:]):
            row_data[fields[j].name] = val
        data.append(row_data)


    return JsonResponse({'status': 'success', 'data': data})


@api_view(['GET'])
def system_dictionary_export_view(request, pk):
    """Generates a professional Excel data dictionary for the entire system."""
    try:
        user, err = require_auth(request)
        if err: return err

        system = get_object_or_404(System, id=pk)
        ok, res = check_system_permission(user, pk, 'read')
        if not ok: return res

        tables = list(SystemTable.objects.filter(system=system, is_deleted=False).order_by('name'))

        wb = openpyxl.Workbook()

        # ── COLORES ──────────────────────────────────────────────────────────
        C_TITLE_BG   = "1E3A8A"   # Azul muy oscuro
        C_HEADER_BG  = "2563EB"   # Azul primario
        C_TABLE_BG   = "1F2937"   # Gris casi negro
        C_FIELD_HDR  = "3B82F6"   # Azul claro
        C_ALT_ROW    = "F0F7FF"   # Azul muy tenue para filas alternas
        C_DESC_BG    = "F9FAFB"   # Gris clarísimo
        C_WHITE      = "FFFFFF"
        C_LIGHT_GRAY = "F3F4F6"
        C_GRAY_TXT   = "374151"

        def hdr_font(color=C_WHITE, size=10, bold=True):
            return Font(bold=bold, size=size, color=color)

        def fill(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        def center():
            return Alignment(horizontal="center", vertical="center")

        # ── HOJA 1: RESUMEN DEL SISTEMA ──────────────────────────────────────
        ws_info = wb.active
        ws_info.title = "Resumen del Sistema"

        # Fila 1 – Título principal
        ws_info.merge_cells('A1:D1')
        ws_info['A1'] = "DATIUM – DICCIONARIO DE DATOS PROFESIONAL"
        ws_info['A1'].font = Font(bold=True, size=14, color=C_WHITE)
        ws_info['A1'].fill = fill(C_TITLE_BG)
        ws_info['A1'].alignment = center()
        ws_info.row_dimensions[1].height = 38

        # Fila 2 – vacía de separación
        ws_info.append(["", "", "", ""])

        # Fila 3 – Sección info general
        ws_info.append(["INFORMACIÓN DEL SISTEMA", "", "", ""])
        ws_info.merge_cells('A3:D3')
        ws_info['A3'].font = hdr_font(size=11)
        ws_info['A3'].fill = fill(C_TABLE_BG)
        ws_info['A3'].alignment = center()
        ws_info.row_dimensions[3].height = 22

        # Filas 4-8 – datos del sistema
        info_rows = [
            ("Nombre del Sistema",     system.name),
            ("Descripción",            system.description or "Sin descripción registrada"),
            ("Propietario",            system.owner.email),
            ("Fecha de Generación",    timezone.now().strftime("%Y-%m-%d %H:%M")),
            ("Total de Tablas",        len(tables)),
        ]
        for idx, (prop, val) in enumerate(info_rows, start=4):
            ws_info.append([prop, val, "", ""])
            ws_info.merge_cells(f'B{idx}:D{idx}')
            ws_info.cell(row=idx, column=1).font = Font(bold=True, size=10, color=C_GRAY_TXT)
            ws_info.cell(row=idx, column=1).fill = fill(C_LIGHT_GRAY)
            ws_info.cell(row=idx, column=2).font = Font(size=10)

        # Filas 9-10 – espaciado
        ws_info.append(["", "", "", ""])
        ws_info.append(["", "", "", ""])

        # Fila 11 – Cabecera del índice
        ws_info.append(["ÍNDICE DE TABLAS", "", "", ""])
        ws_info.merge_cells('A11:D11')
        ws_info['A11'].font = hdr_font(size=12)
        ws_info['A11'].fill = fill(C_TABLE_BG)
        ws_info['A11'].alignment = center()
        ws_info.row_dimensions[11].height = 24

        # Fila 12 – Cabecera columnas del índice
        index_headers = ["Nombre de la Tabla", "Descripción", "Nro. Columnas", "Nro. Registros"]
        ws_info.append(index_headers)
        for col_idx in range(1, 5):
            c = ws_info.cell(row=12, column=col_idx)
            c.font = hdr_font()
            c.fill = fill(C_HEADER_BG)
            c.alignment = center()
        ws_info.row_dimensions[12].height = 18

        for row_num, t in enumerate(tables, start=13):
            n_fields  = SystemField.objects.filter(table=t).count()
            n_records = SystemRecord.objects.filter(table=t, is_deleted=False).count()
            ws_info.append([t.name, t.description or "Sin descripción", n_fields, n_records])
            # Filas alternas
            bg = C_ALT_ROW if row_num % 2 == 0 else C_WHITE
            for col_idx in range(1, 5):
                c = ws_info.cell(row=row_num, column=col_idx)
                c.font = Font(size=10)
                c.fill = fill(bg)

        # Anchos fijos hoja 1
        ws_info.column_dimensions['A'].width = 28
        ws_info.column_dimensions['B'].width = 42
        ws_info.column_dimensions['C'].width = 16
        ws_info.column_dimensions['D'].width = 16

        # ── HOJA 2: DICCIONARIO DETALLADO ────────────────────────────────────
        ws_dict = wb.create_sheet(title="Diccionario de Datos")

        # Fila 1 – Título
        ws_dict.merge_cells('A1:F1')
        ws_dict['A1'] = "DICCIONARIO DE DATOS – DETALLE POR TABLA"
        ws_dict['A1'].font = Font(bold=True, size=14, color=C_WHITE)
        ws_dict['A1'].fill = fill(C_TITLE_BG)
        ws_dict['A1'].alignment = center()
        ws_dict.row_dimensions[1].height = 38

        current_row = 2

        FIELD_HEADERS = [
            "Nombre del Campo",
            "Tipo de Dato",
            "Llave Primaria (PK)",
            "Único (Unique)",
            "Obligatorio (Not Null)",
            "Relación / Tabla Destino",
        ]

        for table in tables:
            # Espacio entre tablas
            ws_dict.append([""] * 6)
            current_row += 1

            # ── Cabecera de tabla ────────────────────────────────────────
            ws_dict.append([f"  TABLA: {table.name.upper()}"] + [""] * 5)
            ws_dict.merge_cells(start_row=current_row, start_column=1,
                                end_row=current_row, end_column=6)
            cell = ws_dict.cell(row=current_row, column=1)
            cell.font = Font(bold=True, size=12, color=C_WHITE)
            cell.fill = fill(C_TABLE_BG)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_dict.row_dimensions[current_row].height = 26
            current_row += 1

            # ── Descripción de tabla ─────────────────────────────────────
            ws_dict.append([f"  {table.description or 'Sin descripción registrada'}"] + [""] * 5)
            ws_dict.merge_cells(start_row=current_row, start_column=1,
                                end_row=current_row, end_column=6)
            cell = ws_dict.cell(row=current_row, column=1)
            cell.font = Font(italic=True, size=10, color=C_GRAY_TXT)
            cell.fill = fill(C_DESC_BG)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_dict.row_dimensions[current_row].height = 18
            current_row += 1

            # ── Cabecera de campos ───────────────────────────────────────
            ws_dict.append(FIELD_HEADERS)
            for col_idx in range(1, 7):
                c = ws_dict.cell(row=current_row, column=col_idx)
                c.font = hdr_font(size=9)
                c.fill = fill(C_FIELD_HDR)
                c.alignment = center()
            ws_dict.row_dimensions[current_row].height = 18
            current_row += 1

            # ── Filas de campos ──────────────────────────────────────────
            fields = list(SystemField.objects.filter(table=table).order_by('order_index'))
            if not fields:
                ws_dict.append(["  (Sin campos definidos)"] + [""] * 5)
                ws_dict.merge_cells(start_row=current_row, start_column=1,
                                    end_row=current_row, end_column=6)
                ws_dict.cell(row=current_row, column=1).font = Font(italic=True, size=9, color="9CA3AF")
                current_row += 1
            else:
                for f_idx, f in enumerate(fields):
                    rel_target = f.related_table.name if f.related_table else ""
                    ws_dict.append([
                        f.name,
                        f.type.upper(),
                        "SÍ" if f.is_primary_key else "—",
                        "SÍ" if f.is_unique else "—",
                        "SÍ" if f.required else "NO (Opcional)",
                        rel_target,
                    ])
                    bg = C_ALT_ROW if f_idx % 2 == 0 else C_WHITE
                    for col_idx in range(1, 7):
                        c = ws_dict.cell(row=current_row, column=col_idx)
                        c.font = Font(size=10)
                        c.fill = fill(bg)
                        if col_idx in (3, 4, 5):
                            c.alignment = center()
                    # Resaltar en ámbar si es PK
                    if f.is_primary_key:
                        ws_dict.cell(row=current_row, column=3).font = Font(bold=True, size=10, color="D97706")
                    current_row += 1

        # Anchos fijos hoja 2
        col_widths = [28, 14, 20, 16, 22, 26]
        for col_idx, w in enumerate(col_widths, start=1):
            ws_dict.column_dimensions[get_column_letter(col_idx)].width = w

        # ── RESPUESTA HTTP ────────────────────────────────────────────────────
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        now_str = timezone.now().strftime("%Y%m%d_%H%M")
        import re
        safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '', system.name.replace(' ', '_')) or "Sistema"
        response['Content-Disposition'] = f'attachment; filename="Diccionario_{safe_name}_{now_str}.xlsx"'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.read())
        return response
    except Exception as e:
        import traceback
        logger.error(f"Error generating Excel dictionary: {str(e)}\n{traceback.format_exc()}")
        return JsonResponse({'error': 'Error crítico al generar el diccionario: ' + str(e)}, status=500)
@api_view(['GET'])
def admin_users_page_view(request):
    user, err = require_admin(request)
    if err: return err
    return render(request, 'admin_users.html', {'user': user})

# ═══════════════════════════════════════════
# ADMIN USER MANAGEMENT (JSON API)
# ═══════════════════════════════════════════

@api_view(['GET'])
def admin_stats_view(request):
    user, err = require_admin(request)
    if err:
        return err
    return Response({
        'users': User.objects.count(),
        'systems': System.objects.filter(is_deleted=False).count(),
    })


@api_view(['GET'])
def admin_users_list_view(request):
    user, err = require_admin(request)
    if err: return err
    
    users = User.objects.all().order_by('-created_at')
    data = []
    for u in users:
        data.append({
            'id': u.id,
            'email': u.email,
            'name': u.name or '',
            'role': u.role,
            'is_suspended': u.is_suspended,
            'createdAt': u.created_at.isoformat()
        })
    return Response(data)

@api_view(['POST'])
def admin_user_action_view(request, pk):
    user, err = require_admin(request)
    if err: return err
    
    try:
        target = User.objects.get(id=pk)
    except User.DoesNotExist:
        return Response({'error': 'Usuario no encontrado'}, status=404)
        
    action = request.data.get('action') # suspend, activate, block_ip
    
    if action == 'suspend':
        target.is_suspended = True
        target.save()
        return Response({'ok': True, 'msg': 'Usuario suspendido'})
    elif action == 'activate':
        target.is_suspended = False
        target.save()
        return Response({'ok': True, 'msg': 'Usuario activado'})
    elif action == 'block_ip':
        ip = request.data.get('ip')
        if ip:
            BlockedIP.objects.get_or_create(ip_address=ip, defaults={'reason': 'Bloqueado por administrador'})
            return Response({'ok': True, 'msg': f'IP {ip} bloqueada'})
            
    return Response({'error': 'Acción inválida'}, status=400)

@api_view(['GET', 'DELETE'])
def admin_blocked_ips_view(request):
    user, err = require_admin(request)
    if err: return err
    if request.method == 'GET':
        ips = BlockedIP.objects.all().order_by('-created_at')
        return Response([{'id': i.id, 'ip': i.ip_address, 'reason': i.reason or '', 'createdAt': i.created_at.isoformat()} for i in ips])
    else:
        ip = request.GET.get('ip')
        if ip:
            BlockedIP.objects.filter(ip_address=ip).delete()
            return Response({'ok': True})
        return Response({'error': 'IP requerida'}, status=400)


@api_view(['GET', 'POST'])
def admin_tyc_view(request):
    user, err = require_admin(request)
    if err: return err
    if request.method == 'GET':
        content_setting, _ = AppSetting.objects.get_or_create(key='terms_content', defaults={'value': '<h1>Términos y Condiciones</h1><p>Acepta para continuar.</p>'})
        version_setting, _ = AppSetting.objects.get_or_create(key='terms_version', defaults={'value': '1'})
        return Response({'content': content_setting.value, 'version': version_setting.value})
    content = request.data.get('content')
    if content:
        content_setting, _ = AppSetting.objects.get_or_create(key='terms_content')
        content_setting.value = content
        content_setting.save()
        version_setting, _ = AppSetting.objects.get_or_create(key='terms_version', defaults={'value': '0'})
        new_version = int(version_setting.value or 0) + 1
        version_setting.value = str(new_version)
        version_setting.save()
        return Response({'ok': True, 'version': new_version})
    return Response({'error': 'Contenido requerido'}, status=400)


@api_view(['GET'])
def admin_payments_view(request):
    user, err = require_admin(request)
    if err: return err
    payments = Payment.objects.select_related('user').all().order_by('-created_at')
    return Response([{
        'id': p.id, 'user': p.user.email if p.user else '', 'amount': float(p.amount),
        'status': p.status, 'createdAt': p.created_at.isoformat()
    } for p in payments])


@api_view(['POST'])
def process_payment_view(request):
    user, err = require_auth(request)
    if err: return err
    return Response({'ok': True, 'message': 'Pago procesado (Simulación)'})

@api_view(['GET'])
def admin_trash_systems_view(request):
    user, err = require_admin(request)
    if err: return err
    systems = System.objects.filter(is_deleted=True)
    return Response([serialize_system(s) for s in systems])

@api_view(['POST'])
def admin_trash_restore_view(request, pk):
    user, err = require_admin(request)
    if err: return err
    system = get_object_or_404(System, id=pk)
    system.is_deleted = False
    system.save()
    return Response({'ok': True})

@api_view(['GET'])
def admin_trash_export_view(request):
    user, err = require_admin(request)
    if err:
        return Response({'error': 'Permiso denegado'}, status=403)
    systems = System.objects.filter(is_deleted=True)
    format_param = request.GET.get('format', 'csv').lower()
    header = ['ID', 'Nombre', 'Descripción', 'Dueño', 'Fecha Creación']
    rows = []
    for s in systems:
        rows.append([s.id, s.name, s.description, s.owner.email, s.created_at.strftime("%Y-%m-%d %H:%M")])
    return generate_export_response(header, rows, "admin_trash_systems", format_param, "Sistemas en Papelera")

@api_view(['GET'])
def admin_export_users_view(request):
    user, err = require_admin(request)
    if err:
        return Response({'error': 'Permiso denegado'}, status=403)
    users = User.objects.all().order_by('-created_at')
    format_param = request.GET.get('format', 'csv').lower()
    header = ['ID', 'Nombre', 'Email', 'Rol', 'Plan', 'Estado', 'Registro']
    rows = []
    for u in users:
        plan_name = u.plan.name if u.plan else 'Gratis'
        st = 'Bloqueado' if u.is_suspended else 'Activo'
        rows.append([u.id, u.name, u.email, u.role, plan_name, st, u.created_at.strftime("%Y-%m-%d")])
    return generate_export_response(header, rows, "admin_users", format_param, "Listado de Usuarios")

@api_view(['GET'])
def admin_export_reports_view(request):
    user, err = require_admin(request)
    if err:
        return Response({'error': 'Permiso denegado'}, status=403)
    reports = UserReport.objects.all().order_by('-created_at')
    format_param = request.GET.get('format', 'csv').lower()
    header = ['ID', 'Título', 'Usuario', 'Estado', 'Fecha']
    rows = []
    for r in reports:
        rows.append([r.id, r.title, r.user.email, r.status, r.created_at.strftime("%Y-%m-%d %H:%M")])
    return generate_export_response(header, rows, "admin_reports", format_param, "Reportes de Usuario")

@api_view(['PUT'])
def update_table_style_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    table = get_object_or_404(SystemTable, id=pk)
    ok, res = check_table_permission(user, table, 'update')
    if not ok:
        return res
    table.custom_style_json = json.dumps(request.data.get('style', {}))
    table.save()
    return Response({'ok': True})

@api_view(['POST'])
def accept_terms_view(request):
    user, err = require_auth(request)
    if err: return err
    version = request.data.get('version', 1)
    user.terms_version_accepted = version
    user.save()
    return Response({'ok': True})

@api_view(['PUT'])
def update_security_settings_view(request):
    user, err = require_auth(request)
    if err: return err
    timeout = request.data.get('session_timeout_minutes')
    if timeout is not None:
        user.session_timeout_minutes = int(timeout)
        user.save()
    return Response({'ok': True})

@api_view(['PUT'])
def reorder_fields_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    table = get_object_or_404(SystemTable, id=pk)
    if table.system.owner_id != user.id:
        return Response({'error': 'Solo el propietario puede reorganizar campos.'}, status=403)
    ok, res = check_system_permission(user, table.system_id, 'read')
    if not ok: return res

    order = request.data.get('order', [])
    for index, field_id in enumerate(order):
        SystemField.objects.filter(id=field_id, table=table).update(order_index=index)
    
    log_action(user, table.system, 'REORGANIZAR_CAMPOS', f'Campos de la tabla "{table.name}" reorganizados')
    return Response({'ok': True})

@api_view(['PUT'])
def reorder_records_view(request, pk):
    user, err = require_auth(request)
    if err: return err
    table = get_object_or_404(SystemTable, id=pk)
    ok, res = check_table_permission(user, table, 'update')
    if not ok: return res

    order = request.data.get('order', [])
    for index, record_id in enumerate(order):
        SystemRecord.objects.filter(id=record_id, table=table).update(order_index=index)
    
    log_action(user, table.system, 'REORGANIZAR_REGISTROS', f'Registros de la tabla "{table.name}" reorganizados')
    return Response({'ok': True})



# ===== ADMIN PLAN/DISCOUNT HOTFIX =====
@api_view(['GET', 'POST'])
def admin_plans_view(request):
    if request.method == 'GET':
        user, err = require_auth(request)
    else:
        user, err = require_admin(request)
    if err: return err
    if request.method == 'GET':
        plans = Plan.objects.all().order_by('price', 'name')
        data = []
        for p in plans:
            promo = {}
            try:
                promo = json.loads(p.features_json or '{}') if (p.features_json or '').strip().startswith('{') else {'features': json.loads(p.features_json or '[]')}
            except Exception:
                promo = {'raw': p.features_json or ''}
            data.append({
                'id': p.id,
                'name': p.name,
                'max_systems': p.max_systems,
                'max_tables_per_system': p.max_tables_per_system,
                'max_records_per_table': p.max_records_per_table,
                'max_fields_per_table': p.max_fields_per_table,
                'max_storage_mb': p.max_storage_mb,
                'price': float(p.price),
                'is_active': p.is_active,
                'has_ai_assistant': p.has_ai_assistant,
                'promo': promo,
            })
        return Response(data)

    payload = request.data
    pid = payload.get('id')
    promo = payload.get('promo') or {}
    features_json = json.dumps(promo, ensure_ascii=False)
    fields = {
        'name': payload.get('name'),
        'max_systems': int(payload.get('max_systems') or 1),
        'max_tables_per_system': int(payload.get('max_tables_per_system') or 3),
        'max_records_per_table': int(payload.get('max_records_per_table') or 50000),
        'max_fields_per_table': int(payload.get('max_fields_per_table') or 200),
        'max_storage_mb': int(payload.get('max_storage_mb') or 1024),
        'price': payload.get('price') or 0,
        'is_active': bool(payload.get('is_active', True)),
        'has_ai_assistant': bool(payload.get('has_ai_assistant', True)),
        'features_json': features_json,
    }
    if pid:
        plan = get_object_or_404(Plan, id=pid)
        for k, v in fields.items():
            setattr(plan, k, v)
        plan.save()
    else:
        plan = Plan.objects.create(**fields)
    return Response({'ok': True, 'id': plan.id})


@api_view(['GET', 'PUT', 'DELETE'])
def admin_plan_detail_view(request, pk):
    user, err = require_admin(request)
    if err: return err
    plan = get_object_or_404(Plan, id=pk)
    if request.method == 'GET':
        promo = {}
        try:
            promo = json.loads(plan.features_json or '{}') if (plan.features_json or '').strip().startswith('{') else {'features': json.loads(plan.features_json or '[]')}
        except Exception:
            promo = {'raw': plan.features_json or ''}
        return Response({
            'id': plan.id,
            'name': plan.name,
            'max_systems': plan.max_systems,
            'max_tables_per_system': plan.max_tables_per_system,
            'max_records_per_table': plan.max_records_per_table,
            'max_fields_per_table': plan.max_fields_per_table,
            'max_storage_mb': plan.max_storage_mb,
            'price': float(plan.price),
            'is_active': plan.is_active,
            'has_ai_assistant': plan.has_ai_assistant,
            'promo': promo,
        })
    if request.method == 'DELETE':
        plan.delete()
        return Response({'ok': True})
    promo = request.data.get('promo')
    for field in ['name','max_systems','max_tables_per_system','max_records_per_table','max_fields_per_table','max_storage_mb','price','is_active','has_ai_assistant']:
        if field in request.data:
            setattr(plan, field, request.data.get(field))
    if promo is not None:
        plan.features_json = json.dumps(promo, ensure_ascii=False)
    plan.save()
    return Response({'ok': True})


@api_view(['GET', 'POST'])
def admin_discounts_view(request):
    user, err = require_admin(request)
    if err: return err
    if request.method == 'POST':
        payload = request.data
        did = payload.get('id')
        code = (payload.get('code') or '').strip().upper()
        percentage = payload.get('percentage') or 0
        if not code:
            return Response({'error': 'C?digo requerido'}, status=400)
        if did:
            disc = get_object_or_404(Discount, id=did)
            disc.code = code
            disc.percentage = percentage
            disc.is_active = bool(payload.get('is_active', True))
            disc.save()
        else:
            disc = Discount.objects.create(code=code, percentage=percentage, is_active=bool(payload.get('is_active', True)))
        return Response({'ok': True, 'id': disc.id})
    discounts = Discount.objects.all().order_by('-created_at')
    return Response([{'id': d.id, 'code': d.code, 'percentage': float(d.percentage), 'is_active': d.is_active} for d in discounts])


@api_view(['PUT', 'DELETE'])
def admin_discount_detail_view(request, pk):
    user, err = require_admin(request)
    if err: return err
    discount = get_object_or_404(Discount, id=pk)
    if request.method == 'DELETE':
        discount.delete()
        return Response({'ok': True})
    discount.code = (request.data.get('code') or discount.code).strip().upper()
    discount.percentage = request.data.get('percentage', discount.percentage)
    discount.is_active = request.data.get('is_active', discount.is_active)
    discount.save()
    return Response({'ok': True})


@api_view(['GET'])
def public_plans_view(request):
    sync_plans_into_db()
    plans = Plan.objects.filter(is_active=True, name__in=['Free', 'Pro', 'Corporate']).order_by('price')
    return Response([{
        'id': p.id,
        'name': p.name,
        'price': float(p.price),
        'max_systems': p.max_systems,
        'max_tables_per_system': p.max_tables_per_system,
        'max_storage_mb': p.max_storage_mb,
        'has_ai_assistant': p.has_ai_assistant,
    } for p in plans])
